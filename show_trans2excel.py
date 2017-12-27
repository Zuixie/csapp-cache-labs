#! /usr/bin/env python3
# -*- coding:utf-8 -*-

'''
this module for csapp cache labs.
read trans result file and show miss result to excel

csapp cache实验PartB部分生成的结果在经过csim处理后，获取矩阵上每个元素的miss情况，
然后次脚本就是将每个元素的miss情况还原到excel中，便于直观比较不同算法直接的结果

step:
1. read file -> data list, get A[0][0] address, B[0][0] address
2. judge this data is A or B, & show compute i,j
3. show in excel
'''

from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import PatternFill

ignore_head = 4
ignore_tail = 2
filename = "trace.ji.f0.csim"
out_excel_filename = "result.xlsx"
matrix_a_addr = int(0)
matrix_b_addr = int(0)
matrix_m = 61 
matrix_n = 67


class Element:
    '''
    store the matrix element info
    '''
    def __init__(self, op='', address=0, miss_count=0):
        ''' init '''
        self.operation = op
        self.address = address
        self.miss_count = miss_count
        self.col = -1
        self.row = -1
        self.index = -1
        self.cache_set = -1
        self.matrix_name = ''


def read_file(filename):
    ''' read file '''
    global matrix_a_addr
    global matrix_b_addr

    ele_list = list()
    with open(filename) as f:
        content = f.readlines()[ignore_head : -1 * ignore_tail]
    content = [x.strip() for x in content]

    for lines in content:
        miss_count = 0
        addr = 0
        if 'miss' in lines:
            miss_count = 1

        templist = lines.split(' ')
        addr = int(templist[1].split(',')[0], 16)
        # matrix_a_addr -> L min address, matrix_b_addr -> S min address
        if templist[0] == 'L' and (matrix_a_addr == 0 or matrix_a_addr > addr):
            matrix_a_addr = addr
        elif templist[0] == 'S' and (matrix_b_addr == 0 or matrix_b_addr > addr):
            matrix_b_addr = addr
        ele_list.append(Element(templist[0], addr, miss_count))

    return ele_list

def location_matrix(element_list):
    '''
    compute the col,row,index with Element
    '''
    for element in element_list:
        addr = element.address
        if addr >= matrix_b_addr:
            element.matrix_name = 'B'
            element.index = int((addr - matrix_b_addr) / 4)
            (element.row, element.col) = divmod(element.index, matrix_n)
        else:
            element.matrix_name = 'A'
            element.index = int((addr - matrix_a_addr) / 4)
            (element.row, element.col) = divmod(element.index, matrix_m)

        element.cache_set = int(element.index / 8) % 32

def show_elements(element_list):
    '''
    show elements,store in excel
    '''
    cell = None
    count_a_miss = 0
    count_b_miss = 0
    wb = Workbook()
    wsa = wb.active
    wsa.title = "matrix A"
    wsb = wb.create_sheet(title="matrix B")

    cell_bg = PatternFill(patternType='solid', fgColor=colors.RED)

    for element in element_list:
        print(element.matrix_name, "i:", element.row, "j:", element.col,
              "index:", element.index, "set:", element.cache_set)
        if element.matrix_name is 'A':
            cell = wsa.cell(column=element.col+1, row=element.row+1,
                            value=str(element.cache_set))
            if element.miss_count > 0:
                count_a_miss = count_a_miss + 1
        else:
            cell = wsb.cell(column=element.col+1, row=element.row+1,
                            value=str(element.cache_set))
            if element.miss_count > 0:
                count_b_miss = count_b_miss + 1
        if element.miss_count > 0:
            cell.fill = cell_bg

    print("A:",count_a_miss, "B:", count_b_miss)
    #wb.save(filename=out_excel_filename)

def main():
    '''main'''
    ele_list = read_file(filename)
    print("matrix_a:", hex(matrix_a_addr), "matrix_b:",
          hex(matrix_b_addr), "op_count:", len(ele_list))

    location_matrix(ele_list)

    show_elements(ele_list)


if __name__ == "__main__":
    main()
